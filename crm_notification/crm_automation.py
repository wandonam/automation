import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

#drvie: 주문서(전처리)
df = pd.read_excel(f'../cmn_data/raw_2024.xlsx')
#cafe24: 고객 -> 회원조회 -> 주문 회원 조회
db_user = pd.read_csv(f'./data/db_order.csv')
#drive: gross management
db_product = pd.read_excel(f'./data/db_product.xlsx')

df = df[df['매출처'] == '자사몰']
df = df[df['결제일'] >= '2024-05-01']

merged_df = df.merge(db_user, left_on='쇼핑몰 주문번호', right_on='최근 주문번호', how='left')
merged_df2 = merged_df.merge(db_product, left_on='옵션코드', right_on='품목코드', how='left')

merged_df2['할인율'] = np.where(merged_df2['20%\n쿠폰'] == 'O', '20%', 
                             np.where(merged_df2['10%\n쿠폰'] == 'O', '10%', '0%'))

crm = pd.DataFrame()
crm['date'] = merged_df2['결제일']
crm['order_no'] = merged_df2['쇼핑몰 주문번호']
crm['order_code'] = merged_df2['옵션코드']
crm['order_prod'] = merged_df2['상품명_x']
crm['user_name'] = merged_df2['이름']
crm['user_tel'] = merged_df2['휴대폰']
crm['period'] = merged_df2['섭취\n기간']
crm['url'] = merged_df2['URL']
crm['scheme'] = merged_df2['할인율']

crm = crm.dropna(subset=['user_tel', 'period', 'url'])
crm['period'] = crm['period'].fillna(0).astype(int)
crm['send'] = crm['date'] + crm['period'].apply(lambda x: timedelta(days=x))
crm['sending_crm'] = crm['send'].apply(lambda x: x + timedelta(days=(5 - x.weekday()) % 7))

today = datetime.today()
this_saturday = today + timedelta((5-today.weekday()) % 7)
this_saturday_f = this_saturday.strftime('%Y-%m-%d')

crm_filtered = crm[crm['sending_crm'] == this_saturday_f]

this_saturday_str = this_saturday.strftime('%y%m%d')

output_path = f'./data/{this_saturday_str}_crm.xlsx'
crm_filtered.to_excel(output_path, index=False, engine='openpyxl')

wb = load_workbook(output_path)
ws = wb.active

date_columns = ['A', 'J', 'K']
date_style = NamedStyle(name='datetime', number_format='YYYY-MM-DD')

for col in date_columns:
    for cell in ws[col][1:]:
        cell.style = date_style

wb.save(output_path)